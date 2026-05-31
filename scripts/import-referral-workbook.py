import json
import os
import re
import hashlib
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

SOURCE = Path(r"C:\Users\daiyu\Downloads\推薦碼大全.xlsx")
OUTPUT = Path(r"C:\Users\daiyu\Desktop\code-share-nextjs\data\referral-data.json")

CATEGORY_IDS = {
    "銀行類": "banking",
    "行動支付類": "mobile-payment",
    "餐飲類": "food",
    "遊戲娛樂類": "game-entertainment",
    "網購類": "shopping",
    "其他APP": "other-app",
}

POPULAR_PLATFORM_NAMES = {
    "Foodpanda",
    "Ubereats",
    "Uber",
    "Line Pay",
    "街口支付",
    "環保集點",
    "蝦皮購物",
    "MOMO",
    "Richart",
}

URL_PATTERN = re.compile(r"https?://[^\s]+")
PHONE_PATTERN = re.compile(r"09\d{8}")


def compact_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slugify(text, fallback_prefix):
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", ascii_text).strip("-").lower()
    digest = hashlib.sha1(text.encode("utf-8")).hexdigest()[:8]
    if slug:
        return f"{slug}-{digest}"
    return f"{fallback_prefix}-{digest}"


def make_id(*parts):
    joined = "|".join(str(part) for part in parts)
    return hashlib.sha1(joined.encode("utf-8")).hexdigest()[:16]


def get_hyperlink(cell):
    if cell.hyperlink is None:
        return None
    target = cell.hyperlink.target or cell.hyperlink.location
    return compact_text(target)


def split_value(raw_value, hyperlink):
    raw = compact_text(raw_value)
    url = hyperlink
    code = None
    display_type = "code"

    url_match = URL_PATTERN.search(raw or "")
    if url_match:
        url = url_match.group(0)
        before_url = compact_text((raw or "")[: url_match.start()])
        if before_url and "點我" not in before_url and "★" not in before_url:
            code = before_url
        display_type = "code_with_link" if code else "link"
    elif hyperlink:
        if raw and "點我" not in raw and "★" not in raw:
            code = raw
            display_type = "code_with_link"
        else:
            display_type = "link"
    else:
        code = raw
        display_type = "code"

    if code:
        code = re.sub(r"\s+", " ", code).strip()

    return code, url, display_type


def quality_flags_for_value(raw, code, url, hyperlink):
    flags = []
    if raw and ("★" in raw or "點我" in raw):
        if hyperlink or url:
            flags.append("placeholder_text_with_hyperlink")
        else:
            flags.append("placeholder_text_without_hyperlink")
    target = " ".join(part for part in [raw or "", code or "", url or ""] if part)
    if PHONE_PATTERN.search(target):
        flags.append("phone_like_value_review_required")
    if raw and "\n" in raw:
        flags.append("multiline_raw_value")
    return flags


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)

    wb = load_workbook(SOURCE, data_only=True)
    categories = []
    platforms = []
    invite_codes = []
    review_issues = []

    for sheet_index, ws in enumerate(wb.worksheets):
        category_name = ws.title
        category_id = CATEGORY_IDS.get(category_name, slugify(category_name, "category"))
        categories.append({
            "id": category_id,
            "name": category_name,
            "sortOrder": sheet_index + 1,
            "sourceSheet": category_name,
        })

        for col_idx in range(2, ws.max_column + 1):
            header = compact_text(ws.cell(row=1, column=col_idx).value)
            if not header or header.startswith("Unnamed"):
                review_issues.append({
                    "type": "unnamed_platform_column",
                    "severity": "warning",
                    "sourceSheet": category_name,
                    "sourceColumn": col_idx,
                })
                continue

            platform_name = re.sub(r"\s+", " ", header).strip()
            platform_id = make_id(category_id, platform_name)
            platform_slug = slugify(platform_name, "platform")
            activity = compact_text(ws.cell(row=2, column=col_idx).value)
            platform_code_ids = []
            platform_flags = []

            if not activity:
                platform_flags.append("missing_activity_description")

            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                raw = compact_text(cell.value)
                hyperlink = get_hyperlink(cell)

                if not raw and not hyperlink:
                    continue

                code, referral_url, display_type = split_value(raw or hyperlink, hyperlink)
                flags = quality_flags_for_value(raw, code, referral_url, hyperlink)

                if "placeholder_text_without_hyperlink" in flags:
                    review_issues.append({
                        "type": "placeholder_without_hyperlink",
                        "severity": "warning",
                        "sourceSheet": category_name,
                        "sourceColumn": col_idx,
                        "sourceRow": row_idx,
                        "platformName": platform_name,
                        "rawValue": raw,
                    })
                    continue

                invite_code_id = make_id(category_id, platform_name, row_idx, raw or referral_url)
                platform_code_ids.append(invite_code_id)

                for flag in flags:
                    severity = "review" if flag == "phone_like_value_review_required" else "info"
                    review_issues.append({
                        "type": flag,
                        "severity": severity,
                        "sourceSheet": category_name,
                        "sourceColumn": col_idx,
                        "sourceRow": row_idx,
                        "platformName": platform_name,
                        "inviteCodeId": invite_code_id,
                        "rawValue": raw,
                    })

                invite_codes.append({
                    "id": invite_code_id,
                    "platformId": platform_id,
                    "categoryId": category_id,
                    "code": code,
                    "referralUrl": referral_url,
                    "rawValue": raw or referral_url,
                    "displayType": display_type,
                    "status": "active",
                    "verificationStatus": "unverified",
                    "sourceType": "seed_import",
                    "usageCount": 0,
                    "reportCount": 0,
                    "lastUsedAt": None,
                    "lastVerifiedAt": None,
                    "expiresAt": None,
                    "sourceSheet": category_name,
                    "sourceColumn": col_idx,
                    "sourceRow": row_idx,
                    "qualityFlags": flags,
                })

            code_count = len(platform_code_ids)
            if code_count == 0:
                platform_flags.append("no_seed_codes")

            platforms.append({
                "id": platform_id,
                "categoryId": category_id,
                "name": platform_name,
                "slug": platform_slug,
                "activityDescription": activity,
                "status": "active" if code_count > 0 else "wanted",
                "acceptsPlainCode": any(c["platformId"] == platform_id and c["code"] for c in invite_codes),
                "acceptsReferralUrl": any(c["platformId"] == platform_id and c["referralUrl"] for c in invite_codes),
                "isPopular": platform_name in POPULAR_PLATFORM_NAMES,
                "codeCount": code_count,
                "sourceSheet": category_name,
                "sourceColumn": col_idx,
                "qualityFlags": platform_flags,
            })

    duplicate_map = {}
    for item in invite_codes:
        key = item.get("referralUrl") or item.get("code") or item.get("rawValue")
        if not key:
            continue
        duplicate_map.setdefault(key, []).append(item["id"])

    for value, ids in duplicate_map.items():
        if len(ids) > 1:
            review_issues.append({
                "type": "duplicate_value",
                "severity": "info",
                "value": value,
                "inviteCodeIds": ids,
            })
            id_set = set(ids)
            for item in invite_codes:
                if item["id"] in id_set and "duplicate_value" not in item["qualityFlags"]:
                    item["qualityFlags"].append("duplicate_value")

    data = {
        "metadata": {
            "sourceFileName": SOURCE.name,
            "importedAt": datetime.now(timezone.utc).isoformat(),
            "version": 1,
            "counts": {
                "categories": len(categories),
                "platforms": len(platforms),
                "activePlatforms": sum(1 for p in platforms if p["status"] == "active"),
                "wantedPlatforms": sum(1 for p in platforms if p["status"] == "wanted"),
                "inviteCodes": len(invite_codes),
                "referralUrlRecords": sum(1 for c in invite_codes if c["referralUrl"]),
                "plainCodeRecords": sum(1 for c in invite_codes if c["code"] and not c["referralUrl"]),
                "reviewIssues": len(review_issues),
            },
        },
        "categories": categories,
        "platforms": platforms,
        "inviteCodes": invite_codes,
        "reviewIssues": review_issues,
    }

    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(data["metadata"]["counts"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
